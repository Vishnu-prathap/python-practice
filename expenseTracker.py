import xlsxwriter
import openpyxl  # Use openpyxl for appending data

def createExpensesSheet():
    """Creates an Excel sheet with column headers if it doesn't exist."""
    try:
        workbook = xlsxwriter.Workbook('mastersExpenses.xlsx')
        worksheet = workbook.add_worksheet()

        # Adding headers
        headers = ['Date', 'Expense', 'Amount', 'Category', 'Payment Mode', 'Description']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header)  # Writing headers in the first row

        workbook.close()
        print("Excel file created successfully!")
    except Exception as e:
        print(f"Error creating file: {e}")

def addExpense(date, expense, amount, category, paymentMode, description):
    """Appends an expense to the existing Excel sheet."""
    try:
        # Open existing workbook
        workbook = openpyxl.load_workbook('mastersExpenses.xlsx')
        worksheet = workbook.active

        # Find the next empty row
        next_row = worksheet.max_row + 1

        # Data to be inserted
        data = [date, expense, amount, category, paymentMode, description]

        # Append data to the sheet
        worksheet.append(data)

        # Save workbook
        workbook.save('mastersExpenses.xlsx')
        workbook.close()
        print("Expense added successfully!")
    except Exception as e:
        print(f"Error adding expense: {e}")

def main():
    """Main function to handle user input and add expenses."""
    createExpensesSheet()  # Ensure the file exists with headers

    while True:
        date = input('Enter the date of the expense (YYYY-MM-DD): ')
        expense = input('Enter the expense: ')
        amount = input('Enter the amount: ')
        category = input('Enter the category: ')        
        paymentMode = input('Enter the payment mode: ')
        description = input('Enter the description: ')

        addExpense(date, expense, amount, category, paymentMode, description)

        more = input("Do you want to add another expense? (yes/no): ").strip().lower()
        if more != 'yes':
            print("Exiting expense tracker.")
            break

if __name__ == "__main__":
    main()
